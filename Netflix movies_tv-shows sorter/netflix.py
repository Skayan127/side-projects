





# import requests

# url = input('https://www.netflix.com/title/70195800')
# html_output_name = input('netflix2')

# req = requests.get(url, 'html.parser')

# with open(html_output_name, 'w') as f:
#     f.write(req.text)
#     f.close()
from os import link
import requests
import re
import time
import html

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font


start = time.time()

wb = Workbook()
dest_filename = "netflix2.xlsx"

# wb = load_workbook("tim.xlsx")
ws1 = wb.active
ws1.title = "Movies on Netflix"

with open ("netflix_mylist.txt", "r") as myfile:
    data=myfile.readlines()
    data=data[0]
    data=html.unescape(data)
# data1='"""'
# data2='"""'
# data=input("Paste the copied page source code here: ")
# data=data1+data+data2
# data=html.unescape(data)

    # print(type(data)) # "data" is a single, long string

titleName = re.findall(r'"fallback-text">(.*?)<', data)         # List of 50 elements containing the names of the Netflix titles
titleNumber = re.findall(r'watch/(.*?)\?', data)                # List of 50 elements containing the unique URL numbers (not complete links)
titleCover = re.findall(r'padded-container" src="(.*?)"', data) # List of 50 elements containing the complete URLS of the Netflix title covers

# print(len(titleName))
# print(titleName)

# print(len(titleNumber))
# print(titleNumber)

# print(titleCover[0])
print("\n")
titleNumberURL=[]
for i in range(len(titleNumber)):
    titleNumberURL.append("https://www.netflix.com/se-en/title/"+titleNumber[i])    # List of 50 complete URLs of the Netflix titles


# r = requests.get('https://www.netflix.com/us-en/title/60031262') #Suits 9 seasons, 134 episodes (12E,7*16E,10E)
    # texten=r.text

tvNamelist=[]
tvSeasons=[]
tvYear=[]
tvSynopsis=[]
tvGenre=[]
tvStarring=[]
tvMaturity=[]
tvMoodTag=[]
tvHook=[]
tvEpisodes=[]
tvEpisodeLength=[]


movieNamelist=[]
movieDuration=[]
movieYear=[]
movieSynopsis=[]
movieGenre=[]
movieStarring=[]
movieMaturiy=[]
movieMoodTag=[]
movieDirector=[]
movieHook=[]

for i in range(len(titleName)):
    result = requests.get(titleNumberURL[i])
    linkHTML = result.content.decode()# read_link.text
    linkHTML=html.unescape(linkHTML)

    if 'seasonCount":' in linkHTML:
        tvNamelist.append(titleName[i])
        tvSeasons.append(re.findall(r'"seasonCount":(.*?),', linkHTML)[0])
        tvYear.append(re.findall(r'"item-year">(.*?)<', linkHTML)[0])
        tvSynopsis.append(re.findall(r'info-synopsis">(.*?)<', linkHTML)[0])
        tvGenre.append(re.findall(r'"item-genre">(.*?)<', linkHTML)[0])
        
        if '"info-starring">' in linkHTML:
            tvStarring.append(re.findall(r'"info-starring">(.*?)<', linkHTML))
        else:
            tvStarring.append(["NA"])

        tvMaturity.append(re.findall(r'maturity-number">(.*?) <', linkHTML)[0])
        tvMoodTag.append(re.findall(r'item-mood-tag">(.*?) <', linkHTML))
        tvEpisodes.append(re.findall(r'episodeNum":(.*?),', linkHTML))
        tvEpisodeLength.append(re.findall(r'episode-runtime">(.*?)m<', linkHTML))
        
        if '"hook-text">' in linkHTML:
            tvHook.append(re.findall(r'"hook-text">(.*?)<', linkHTML)[0])
        else:
            tvHook.append("NA")
    else:
        movieNamelist.append(titleName[i])
        # movieDuration.append(re.findall(r'duration">(.*?)<', linkHTML)[0])
        a = re.findall(r'duration">(.*?)<', linkHTML)[0]
        # print(a)
        if "h" in a and "m" in a:
            movieDuration.append(int(re.findall(r'(.*?)h', a)[0])*60 + int(re.findall(r'\d+h (.*?)m', a)[0]))
        elif "h" in a and "m" not in a:
            movieDuration.append(int(re.findall(r'(.*?)h', a)[0])*60)
        elif "h" not in a:
            movieDuration.append(int(re.findall(r'\d+h (.*?)m', a)[0]))
    
            # break # movieDuration.append(int(a[:-1]))
            



        movieYear.append(re.findall(r'"item-year">(.*?)<', linkHTML)[0])
        movieSynopsis.append(re.findall(r'info-synopsis">(.*?)<', linkHTML)[0])
        movieGenre.append(re.findall(r'"item-genre">(.*?)<', linkHTML)[0])

        if '"info-starring">' in linkHTML:
            movieStarring.append(re.findall(r'"info-starring">(.*?)<', linkHTML))
        else:
            movieStarring.append(["NA"])

        movieMaturiy.append(re.findall(r'maturity-number">(.*?) <', linkHTML)[0])
        movieMoodTag.append(re.findall(r'item-mood-tag">(.*?) <', linkHTML))
        movieDirector.append(re.findall(r'"director":\[{"@type":"Person","name":"(.*?)"', linkHTML))
        if '"hook-text">' in linkHTML:
            movieHook.append(re.findall(r'"hook-text">(.*?)<', linkHTML)[0])
        else:
            movieHook.append("NA")



    time.sleep(1)


movieStarring2=[]

for i in range(len(movieStarring)):
    movieStarring2.append(movieStarring[i][0])
movieStarring=movieStarring2


tvStarring2=[]

for i in range(len(tvStarring)):
    tvStarring2.append(tvStarring[i][0])
tvStarring=tvStarring2

tvTotalEpisodes=[len(show) for show in tvEpisodes]

tvTotalLength=[]
for i in range(len(tvEpisodeLength)):
    total = 0
    for ele in range(0,len(tvEpisodeLength[i])):
        total = total + int(tvEpisodeLength[i][ele])
    tvTotalLength.append(round(total/60,2))



movieDirector2=[]

for i in range(len(movieDirector)):
    movieDirector2.append(movieDirector[i][0])
movieDirector=movieDirector2





if len(movieNamelist) != 0:
    movieListLength = list(range(1,len(movieNamelist)+1))

if len(tvNamelist) != 0:
    tvListLength = list(range(1,len(tvNamelist)+1))


#movieTitle=['Ex Machina', 'tick, tick...BOOM!', 'The Harder They Fall', 'Bohemian Rhapsody', 'Catch Me If You Can']
#movieLength=['1h 48m', '1h 55m', '2h 19m', '2h 14m', '2h 21m']
#movieGenre=['Sci-Fi Films', 'Musicals', 'Dramas', 'LGBTQ Films', 'Films Based on Books']

all=[movieListLength,
    movieNamelist,
    movieDuration,
    movieYear,
    movieGenre,
    movieMaturiy,
    movieStarring,
    movieDirector,
    movieHook,
    movieSynopsis,
    ]

data = {
    "Movies": {
        "Title": 65,
        "Duration": 65,
        "Year": 78,
        "Genre": 65,
        "Maturity": 78,
        "Starring": 78,
        "Director": 65,
        "Hook": 65,
        "Synopsis": 78,
    },
    "TV-shows": {
        "Title": 65,
        "Year": 78,
        "Seasons": 65,
        "Total episodes": 65,
        "Hours to watch": 65,
        "Genre": 65,
        "Maturity": 78,
        "Starring": 78,
        "Hook": 65,
        # "Director": 65,
        "Synopsis": 78,
    },


}
print("Length of episodes of each TV-show")
print(tvEpisodeLength)
print("Total length of all episodes of every TV-show")
print(tvTotalLength)


# print(all)
print(len(all))
print(all[0][0])

headings = ["Number"] + list(data["Movies"].keys())
ws1.append(headings)
for col in range(0,len(all)):    
    for row in range(0,len(movieNamelist)):
        _ = ws1.cell(column=col+1, row=row+2, value=all[col][row])


all2=[tvListLength,
    tvNamelist,
    tvYear,
    tvSeasons,
    tvTotalEpisodes,
    tvTotalLength,
    tvGenre,
    tvMaturity,
    tvStarring,
    tvHook,
    tvSynopsis,
    ]



ws2 = wb.create_sheet(title="TV-shows on Netflix")

headings = ["Number"] + list(data["TV-shows"].keys())
ws2.append(headings)
for col in range(0,len(all2)):    
    for row in range(0,len(tvNamelist)):
        _ = ws2.cell(column=col+1, row=row+2, value=all2[col][row])



wb.save(filename = dest_filename)


print("\n")
print("Names of movies and TV-shows:")
print(len(movieNamelist))
print(movieNamelist)
print(len(tvNamelist))
print(tvNamelist)


# print("\n")
# print("Length of movies and number of seasons for each TV-show:")
# print(movieDuration)
# print(tvSeasons)


# # print("\n")
# # print("Release date (year) of movies and TV-shows:")
# # print(tvYear)
# # print(movieYear)



# print("\n")
# print("Synopsis of movies and TV-shows:")
# print(movieSynopsis)
# print(tvSynopsis)


# print("\n")
# print("Genres of movies and TV-shows:")
# print(movieGenre)
# print(tvGenre)


# # print("\n")
# # print("Stars of movies and TV-shows:")
# # print(movieStarring)
# # print(tvStarring)


# print("\n")
# print("Recommended age for movies and TV-shows:")
# print(movieMaturiy)
# print(tvMaturity)

# # print("\n")
# # print("Mood tag for movies and TV-shows:")
# # print(movieMoodTag)
# # print(tvMoodTag)


print("\n")
print("Movie director:")
print(movieDirector)


# print("\n")
# print("Hook")
# print(movieHook)
# print(tvHook)


print("\n")
print("Episodes")
print(tvEpisodes)

print("\n")
print(tvTotalEpisodes)

end = time.time()
print(f"Runtime of the program is {end - start}")


print("\n")
print("\n")



# https://openpyxl.readthedocs.io/en/stable/usage.html#write-a-workbook



# read_link = requests.get(titleNumberURL[0])

# linkHTML=read_link.text

# print(linkHTML)



#from selenium import webdriver

#browser = webdriver.Chrome()
#browser.get("https://vecka.nu/")
#browser.get("https://www.netflix.com/title/70195800")
#html_source = browser.page_source
#print(html_source)

    # r = requests.get('https://www.netflix.com/se-en/title/60031262') #Suits 9 seasons, 134 episodes (12E,7*16E,10E)
    # texten=r.text


#print(r.text)

#print(type(r.text))
#print(len(r.text))

#numberOfSeasons = re.findall(r'"seasonCount":(.*?),', texten)


    # length = re.findall(r'duration">(.*?)<', texten)

    # print(length[0])

# result = texten.split()
# print(result)

# for i in range(5):
#     print(r.text[i])







# 81256675

# "fallback-text"	#namnet på filmen/serien

# a href="/watch/	#unika numret för filmen/serien


# Används till URL för varje enskild titel (tv-serie elr film)

# ,"episodeCount":

# "hook-text">

# more-details-item item-genres"

# Both:


# **TV-shows:**
# Episode x of Season y
# item-year"> -- DONE
# ,"seasonCount": -- DONE
# a="title-info-synopsis"> -- DONE
# "item-genre"> -- DONE
# "info-starring"> -- DONE
# maturity-number"> -- DONE
# item-mood-tag"> -- DONE

# **Movies:**
# item-year"> -- DONE
# duration">     </span -- DONE
# a="title-info-synopsis"> -- DONE
# "item-genre"> -- DONE
# "info-starring"> -- DONE
# maturity-number"> -- DONE
# item-mood-tag"> -- DONE


# boxart-image boxart-image-in-padded-container"  #cover


# <div class="galleryContent"> ==$0

# När man kopierat html-texten från Netflix "Min lista", lagt in den i en txt-fil (och kanske gett den ett lämpligt namn) är det klart. Eventuellt (innebär extra arbete oavsett egentligen) kan användaren bara kopiera HTML-texten och ladda upp den i skrivfält på sidan (istället för att först spara i en fil and stuff). Kopierar bara raden <div class="galleryContent">

# Nu blev jag plötsligt fundersam. Om "Min lista" är riktigt långt, kan det vara så att alla titlar inte dyker upp i Inspect element om man inte skrollar till sista titeln i listan? Isåfall blir det ju jobbigt...

# Programmet ska kunna läsa in filen (i mitt fall (skrivande stund 03.01 27/11 2021) är det 50 titlar i "Min lista". 51 egentligen, men Once Upon A Time in Hollywood togs bort efter 23:59 dagen innan). Den ska kunna extrahera namnet på varje titel (tv-serie elr film) samt kopiera dess motsvarande siffra som tillhör titelns egna länk, exempelvis kopiera numret 81256675 som ska finnas i https://www.netflix.com/se-en/title/81256675. Därefter är det regex som gäller. Ska förmodligen vara lite if-satser, ifall titeln för varje URL är en serie eller film. Eftersom det är lite skillnad på vilken typ av data som ska samlas för de två typerna. Att skilja dem åt kan man exempelvis testa med att kontrollera ifall URL:ens HTML har ","seasonCount": i sitt innehåll. Ja? Det är en serie. Nej? Film isåfall.
# Sen kan man sortera dessa i sina olika kategorier och sedan kanske automatiskt skrivna in resultatet i en Excel-fil. Vill man gå steget längre kan man göra en simpel hemsida där användaren laddar upp första filen, och sedan sköts allt automatiskt. Kom också på att att en av kolumnerna i Excel-filen kanske kan ha omslaget på titeln?

# Fyfan vad det är mycket som måste göras haha. Men som vi säger på Flyg: det blir kul, och kul är roligt!